import IPython
import tensorflow as tf
import acme
from acme import environment_loop
from acme import specs
from acme import wrappers
from acme.tf import networks
from acme.tf import utils as tf2_utils
from acme.utils import loggers
import numpy as np
import sonnet as snt
import gym
from envs.environment import Environment
from acme import core

# Imports required for visualization
import pyvirtualdisplay
import imageio
import base64

import copy
from typing import Optional

from acme import datasets
from acme.adders import reverb as adders
from acme.agents import agent
from acme.tf import savers as tf2_savers
from acme.tf import utils as tf2_utils
from acme.utils import loggers
import reverb
import trfl

import time
from typing import Dict, List

#import acme
from acme.agents.tf.dqn.learning import DQNLearner
from acme.tf import losses
from acme.utils import counting
import tensorflow as tf

"""Generic actor implementation, using TensorFlow and Sonnet."""

from typing import Optional, Tuple

from acme import adders as adders2
from acme import types
# Internal imports.
from acme.tf import variable_utils as tf2_variable_utils

import dm_env
import tensorflow_probability as tfp

tfd = tfp.distributions


import operator

import dm_env
from dm_env import specs as specs2
import tree


class EnvironmentLoop2(core.Worker):
  """A simple RL environment loop.
  This takes `Environment` and `Actor` instances and coordinates their
  interaction. Agent is updated if `should_update=True`. This can be used as:
    loop = EnvironmentLoop(environment, actor)
    loop.run(num_episodes)
  A `Counter` instance can optionally be given in order to maintain counts
  between different Acme components. If not given a local Counter will be
  created to maintain counts between calls to the `run` method.
  A `Logger` instance can also be passed in order to control the output of the
  loop. If not given a platform-specific default logger will be used as defined
  by utils.loggers.make_default_logger. A string `label` can be passed to easily
  change the label associated with the default logger; this is ignored if a
  `Logger` instance is given.
  """

  def __init__(
      self,
      environment: dm_env.Environment,
      actor: core.Actor,
      counter: counting.Counter = None,
      logger: loggers.Logger = None,
      should_update: bool = True,
      label: str = 'environment_loop',
  ):
    # Internalize agent and environment.
    self._environment = environment
    self._actor = actor
    self._counter = counter or counting.Counter()
    self._logger = logger or loggers.make_default_logger(label)
    self._should_update = should_update

  def run_episode(self) -> loggers.LoggingData:
    """Run one episode.
    Each episode is a loop which interacts first with the environment to get an
    observation and then give that observation to the agent in order to retrieve
    an action.
    Returns:
      An instance of `loggers.LoggingData`.
    """
    # Reset any counts and start the environment.
    start_time = time.time()
    episode_steps = 0

    # For evaluation, this keeps track of the total undiscounted reward
    # accumulated during the episode.
    episode_return = tree.map_structure(_generate_zeros_from_spec,
                                        self._environment.reward_spec())
    timestep = self._environment.reset()

    # Make the first observation.
    self._actor.observe_first(timestep)

    # Run an episode.
    while not timestep.last():
      # Generate an action from the agent's policy and step the environment.
      action = self._actor._actor.select_action2(timestep.observation, self._environment.action_mask)
      timestep = self._environment.step(action)

      # Have the agent observe the timestep and let the actor update itself.
      self._actor.observe(action, next_timestep=timestep)
      if self._should_update:
        self._actor.update()

      # Book-keeping.
      episode_steps += 1

      # Equivalent to: episode_return += timestep.reward
      # We capture the return value because if timestep.reward is a JAX
      # DeviceArray, episode_return will not be mutated in-place. (In all other
      # cases, the returned episode_return will be the same object as the
      # argument episode_return.)
      episode_return = tree.map_structure(operator.iadd,
                                          episode_return,
                                          timestep.reward)

    # Record counts.
    counts = self._counter.increment(episodes=1, steps=episode_steps)

    # Collect the results and combine with counts.
    steps_per_second = episode_steps / (time.time() - start_time)
    result = {
        'episode_length': episode_steps,
        'episode_return': episode_return,
        'steps_per_second': steps_per_second,
        'jobs_completed': environment.stats["jobs_completed"],
        'makespan': environment.stats["makespan"], 
    }
    result.update(counts)
    return result

  def run(self,
          num_episodes: Optional[int] = None,
          num_steps: Optional[int] = None):
    """Perform the run loop.
    Run the environment loop either for `num_episodes` episodes or for at
    least `num_steps` steps (the last episode is always run until completion,
    so the total number of steps may be slightly more than `num_steps`).
    At least one of these two arguments has to be None.
    Upon termination of an episode a new episode will be started. If the number
    of episodes and the number of steps are not given then this will interact
    with the environment infinitely.
    Args:
      num_episodes: number of episodes to run the loop for.
      num_steps: minimal number of steps to run the loop for.
    Raises:
      ValueError: If both 'num_episodes' and 'num_steps' are not None.
    """

    if not (num_episodes is None or num_steps is None):
      raise ValueError('Either "num_episodes" or "num_steps" should be None.')

    def should_terminate(episode_count: int, step_count: int) -> bool:
      return ((num_episodes is not None and episode_count >= num_episodes) or
              (num_steps is not None and step_count >= num_steps))

    episode_count, step_count = 0, 0
    while not should_terminate(episode_count, step_count):
      result = self.run_episode()
      episode_count += 1
      step_count += result['episode_length']
      # Log the given results.
      self._logger.write(result)
      self._logger._file.flush()
    #self._logger._file.close()
      


def _generate_zeros_from_spec(spec: specs2.Array) -> np.ndarray:
  return np.zeros(spec.shape, spec.dtype)

class FeedForwardActor2(core.Actor):
  """A feed-forward actor.
  An actor based on a feed-forward policy which takes non-batched observations
  and outputs non-batched actions. It also allows adding experiences to replay
  and updating the weights from the policy on the learner.
  """

  def __init__(
      self,
      policy_network: snt.Module,
      adder: Optional[adders2.Adder] = None,
      variable_client: Optional[tf2_variable_utils.VariableClient] = None,
  ):
    """Initializes the actor.
    Args:
      policy_network: the policy to run.
      adder: the adder object to which allows to add experiences to a
        dataset/replay buffer.
      variable_client: object which allows to copy weights from the learner copy
        of the policy to the actor copy (in case they are separate).
    """

    # Store these for later use.
    self._adder = adder
    self._variable_client = variable_client
    self._policy_network = policy_network 

  @tf.function
  def _policy(self, observation: types.NestedTensor, mask: types.NestedTensor) -> types.NestedTensor:
    # Add a dummy batch dimension and as a side effect convert numpy to TF.
    batched_observation = tf2_utils.add_batch_dim(observation)

    # Compute the policy, conditioned on the observation.
    qs = self._policy_network(batched_observation)
    
    qs = qs * tf.cast(mask, dtype=tf.float32)
    # Sample from the policy if it is stochastic.
    action = trfl.epsilon_greedy(qs, epsilon=0.05).sample()

    return action

  def select_action(self, observation: types.NestedArray) -> types.NestedArray:
    # Pass the observation through the policy network.
    action = self._policy(observation, mask)

    # Return a numpy array with squeezed out batch dimension.
    return tf2_utils.to_numpy_squeeze(action)

  def select_action2(self, observation: types.NestedArray, mask: types.NestedArray) -> types.NestedArray:
    # Pass the observation through the policy network.
    action = self._policy(observation, mask)

    # Return a numpy array with squeezed out batch dimension.
    return tf2_utils.to_numpy_squeeze(action)

  def observe_first(self, timestep: dm_env.TimeStep):
    if self._adder:
      self._adder.add_first(timestep)

  def observe(self, action: types.NestedArray, next_timestep: dm_env.TimeStep):
    if self._adder:
      self._adder.add(action, next_timestep)

  def update(self, wait: bool = False):
    if self._variable_client:
      self._variable_client.update(wait)


class AnotherDQN(agent.Agent):
  """DQN2 agent.

  This implements a single-process DQN agent. This is a simple Q-learning
  algorithm that inserts N-step transitions into a replay buffer, and
  periodically updates its policy by sampling these transitions using
  prioritization.
  """

  def __init__(
      self,
      environment_spec: specs.EnvironmentSpec,
      network: snt.Module,
      batch_size: int = 512,
      prefetch_size: int = 4,
      target_update_period: int = 100,
      samples_per_insert: float = 32.0,
      min_replay_size: int = 1000,
      max_replay_size: int = 1000000,
      importance_sampling_exponent: float = 0.2,
      priority_exponent: float = 0.6,
      n_step: int = 5,
      epsilon: Optional[tf.Tensor] = None,
      learning_rate: float = 1e-3,
      discount: float = 0.99,
      logger: loggers.Logger = None,
      checkpoint: bool = True,
      checkpoint_subpath: str = '~/acme/',
      policy_network: Optional[snt.Module] = None,
      #max_gradient_norm: Optional[float] = None,
  ):
    """Initialize the agent.
 
    Args:
      environment_spec: description of the actions, observations, etc.
      network: the online Q network (the one being optimized)
      batch_size: batch size for updates.
      prefetch_size: size to prefetch from replay.
      target_update_period: number of learner steps to perform before updating
        the target networks.
      samples_per_insert: number of samples to take from replay for every insert
        that is made.
      min_replay_size: minimum replay size before updating. This and all
        following arguments are related to dataset construction and will be
        ignored if a dataset argument is passed.
      max_replay_size: maximum replay size.
      importance_sampling_exponent: power to which importance weights are raised
        before normalizing.
      priority_exponent: exponent used in prioritized sampling.
      n_step: number of steps to squash into a single transition.
      epsilon: probability of taking a random action; ignored if a policy
        network is given.
      learning_rate: learning rate for the q-network update.
      discount: discount to use for TD updates.
      logger: logger object to be used by learner.
      checkpoint: boolean indicating whether to checkpoint the learner.
      checkpoint_subpath: directory for the checkpoint.
      policy_network: if given, this will be used as the policy network.
        Otherwise, an epsilon greedy policy using the online Q network will be
        created. Policy network is used in the actor to sample actions.
      max_gradient_norm: used for gradient clipping.
    """

    # Create a replay server to add data to. This uses no limiter behavior in
    # order to allow the Agent interface to handle it.
    max_gradient_norm = None
    replay_table = reverb.Table(
        name=adders.DEFAULT_PRIORITY_TABLE,
        sampler=reverb.selectors.Prioritized(priority_exponent),
        remover=reverb.selectors.Fifo(),
        max_size=max_replay_size,
        rate_limiter=reverb.rate_limiters.MinSize(1),
        signature=adders.NStepTransitionAdder.signature(environment_spec))
    self._server = reverb.Server([replay_table], port=None)

    # The adder is used to insert observations into replay.
    address = f'localhost:{self._server.port}'
    adder = adders.NStepTransitionAdder(
        client=reverb.Client(address),
        n_step=n_step,
        discount=discount)

    # The dataset provides an interface to sample from replay.
    replay_client = reverb.TFClient(address)
    dataset = datasets.make_reverb_dataset(
        server_address=address,
        batch_size=batch_size,
        prefetch_size=prefetch_size)

    # Create epsilon greedy policy network by default.
    if policy_network is None:
    # Use constant 0.05 epsilon greedy policy by default.
      if epsilon is None:
        epsilon = tf.Variable(0.05, trainable=False)
      policy_network = snt.Sequential([
        network,
        #lambda q: trfl.epsilon_greedy(q, epsilon=epsilon).sample()
      ])

    # Create a target network.
    target_network = copy.deepcopy(network)

    # Ensure that we create the variables before proceeding (maybe not needed).
    tf2_utils.create_variables(network, [environment_spec.observations])
    tf2_utils.create_variables(target_network, [environment_spec.observations])

    # Create the actor which defines how we take actions.
    actor = FeedForwardActor2(policy_network, adder)

    # The learner updates the parameters (and initializes them).
    learner = DQNLearner(
        network=network,
        target_network=target_network,
        discount=discount,
        importance_sampling_exponent=importance_sampling_exponent,
        learning_rate=learning_rate,
        target_update_period=target_update_period,
        dataset=dataset,
        replay_client=replay_client,
        #max_gradient_norm=max_gradient_norm,
        logger=logger,
        checkpoint=checkpoint)

    if checkpoint:
      self._checkpointer = tf2_savers.Checkpointer(
          directory=checkpoint_subpath,
          objects_to_save=learner.state,
          subdirectory='dqn_learner3',
          time_delta_minutes=10.)
    else:
      self._checkpointer = None

    super().__init__(
        actor=actor,
        learner=learner,
        min_observations=max(batch_size, min_replay_size),
        observations_per_step=float(batch_size) / samples_per_insert)

  def update(self):
    super().update()
    if self._checkpointer is not None:
      self._checkpointer.save()

from acme import adders as adidas
class RecurrentActor2(core.Actor):
  """A recurrent actor.
  An actor based on a recurrent policy which takes non-batched observations and
  outputs non-batched actions, and keeps track of the recurrent state inside. It
  also allows adding experiences to replay and updating the weights from the
  policy on the learner.
  """

  def __init__(
      self,
      policy_network: snt.RNNCore,
      adder: Optional[adidas.Adder] = None,
      variable_client: Optional[tf2_variable_utils.VariableClient] = None,
      store_recurrent_state: bool = True,
  ):
    """Initializes the actor.
    Args:
      policy_network: the (recurrent) policy to run.
      adder: the adder object to which allows to add experiences to a
        dataset/replay buffer.
      variable_client: object which allows to copy weights from the learner copy
        of the policy to the actor copy (in case they are separate).
      store_recurrent_state: Whether to pass the recurrent state to the adder.
    """
    # Store these for later use.
    self._adder = adder
    self._variable_client = variable_client
    self._network = policy_network
    self._state = None
    self._prev_state = None
    self._store_recurrent_state = store_recurrent_state

  @tf.function
  def _policy(
      self,
      observation: types.NestedTensor,
      state: types.NestedTensor, mask: types.NestedTensor
  ) -> Tuple[types.NestedTensor, types.NestedTensor]:

    # Add a dummy batch dimension and as a side effect convert numpy to TF.
    batched_observation = tf2_utils.add_batch_dim(observation)

    # Compute the policy, conditioned on the observation.
    qvals, new_state = self._network(batched_observation, state)

    # Sample from the policy if it is stochastic.
    action = trfl.epsilon_greedy(qvals, epsilon=0.05, legal_actions_mask=tf.cast(mask, dtype=tf.float32)).sample()
    return action, new_state

  def select_action2(self, observation: types.NestedArray, mask: types.NestedArray) -> types.NestedArray:
    # Initialize the RNN state if necessary.
    if self._state is None:
      self._state = self._network.initial_state(1)

    # Step the recurrent policy forward given the current observation and state.
    policy_output, new_state = self._policy(observation, self._state, mask)
    #counter=0
    #while mask[policy_output]==0 and counter<1:
    #    policy_output, new_state = self._policy(observation, self._state, mask)
    #    counter+=1
    #if counter==1:
    #    print("Valid actions are hard to find here! ->"+str(set(mask)))
    # Bookkeeping of recurrent states for the observe method.
    self._prev_state = self._state
    self._state = new_state

    # Return a numpy array with squeezed out batch dimension.
    return tf2_utils.to_numpy_squeeze(policy_output)


  def select_action(self, observation: types.NestedArray) -> types.NestedArray:
    # Initialize the RNN state if necessary.
    if self._state is None:
      self._state = self._network.initial_state(1)

    # Step the recurrent policy forward given the current observation and state.
    policy_output, new_state = self._policy(observation, self._state)
    
    # Bookkeeping of recurrent states for the observe method.
    self._prev_state = self._state
    self._state = new_state

    # Return a numpy array with squeezed out batch dimension.
    return tf2_utils.to_numpy_squeeze(policy_output)

  def observe_first(self, timestep: dm_env.TimeStep):
    if self._adder:
      self._adder.add_first(timestep)

    # Set the state to None so that we re-initialize at the next policy call.
    self._state = None

  def observe(self, action: types.NestedArray, next_timestep: dm_env.TimeStep):
    if not self._adder:
      return

    if not self._store_recurrent_state:
      self._adder.add(action, next_timestep)
      return

    numpy_state = tf2_utils.to_numpy_squeeze(self._prev_state)
    self._adder.add(action, next_timestep, extras=(numpy_state,))

  def update(self, wait: bool = False):
    if self._variable_client:
      self._variable_client.update(wait)


import functools
import time
from typing import Dict, Iterator, List, Mapping, Union, Optional

from acme import specs
from acme.tf import losses
from acme.tf import networks


Variables = List[np.ndarray]


class R2D2Learner2(acme.Learner, tf2_savers.TFSaveable):
  """R2D2 learner2.
  This is the learning component of the R2D2 agent. It takes a dataset as input
  and implements update functionality to learn from this dataset.
  """

  def __init__(
      self,
      environment_spec: specs.EnvironmentSpec,
      network: Union[networks.RNNCore, snt.RNNCore],
      target_network: Union[networks.RNNCore, snt.RNNCore],
      burn_in_length: int,
      sequence_length: int,
      dataset: tf.data.Dataset,
      reverb_client: Optional[reverb.TFClient] = None,
      counter: counting.Counter = None,
      logger: loggers.Logger = None,
      discount: float = 0.99,
      target_update_period: int = 100,
      importance_sampling_exponent: float = 0.2,
      max_replay_size: int = 1_000_000,
      learning_rate: float = 1e-3,
      # TODO(sergomez): rename to use_core_state for consistency with JAX agent.
      store_lstm_state: bool = True,
      max_priority_weight: float = 0.9,
      n_step: int = 5,
      clip_grad_norm: float = None,
  ):

    if not isinstance(network, networks.RNNCore):
      network.unroll = functools.partial(snt.static_unroll, network)
      target_network.unroll = functools.partial(snt.static_unroll,
                                                target_network)

    # Internalise agent components (replay buffer, networks, optimizer).
    # TODO(b/155086959): Fix type stubs and remove.
    self._iterator: Iterator[reverb.ReplaySample] = iter(dataset)  # pytype: disable=wrong-arg-types
    self._network = network
    self._target_network = target_network
    self._optimizer = snt.optimizers.Adam(learning_rate, epsilon=1e-3)
    self._reverb_client = reverb_client

    # Internalise the hyperparameters.
    self._store_lstm_state = store_lstm_state
    self._burn_in_length = burn_in_length
    self._discount = discount
    self._max_replay_size = max_replay_size
    self._importance_sampling_exponent = importance_sampling_exponent
    self._max_priority_weight = max_priority_weight
    self._target_update_period = target_update_period
    self._num_actions = environment_spec.actions.num_values
    self._sequence_length = sequence_length
    self._n_step = n_step
    self._clip_grad_norm = clip_grad_norm

    if burn_in_length:
      self._burn_in = lambda o, s: self._network.unroll(o, s, burn_in_length)
    else:
      self._burn_in = lambda o, s: (o, s)  # pylint: disable=unnecessary-lambda

    # Learner state.
    self._variables = network.variables
    self._num_steps = tf.Variable(
        0., dtype=tf.float32, trainable=False, name='step')

    # Internalise logging/counting objects.
    self._counter = counting.Counter(counter, 'learner')
    self._logger = logger or loggers.TerminalLogger('learner', time_delta=100.)

    # Do not record timestamps until after the first learning step is done.
    # This is to avoid including the time it takes for actors to come online and
    # fill the replay buffer.
    self._timestamp = None

  @tf.function
  def _step(self) -> Dict[str, tf.Tensor]:

    # Draw a batch of data from replay.
    sample: reverb.ReplaySample = next(self._iterator)

    data = tf2_utils.batch_to_sequence(sample.data)
    observations, actions, rewards, discounts, extra = (data.observation,
                                                        data.action,
                                                        data.reward,
                                                        data.discount,
                                                        data.extras)
    unused_sequence_length, batch_size = actions.shape

    # Get initial state for the LSTM, either from replay or simply use zeros.
    if self._store_lstm_state:
      core_state = tree.map_structure(lambda x: x[0], extra['core_state'])
    else:
      core_state = self._network.initial_state(batch_size)
    target_core_state = tree.map_structure(tf.identity, core_state)

    # Before training, optionally unroll the LSTM for a fixed warmup period.
    burn_in_obs = tree.map_structure(lambda x: x[:self._burn_in_length],
                                     observations)
    _, core_state = self._burn_in(burn_in_obs, core_state)
    _, target_core_state = self._burn_in(burn_in_obs, target_core_state)

    # Don't train on the warmup period.
    observations, actions, rewards, discounts, extra = tree.map_structure(
        lambda x: x[self._burn_in_length:],
        (observations, actions, rewards, discounts, extra))

    with tf.GradientTape() as tape:
      # Unroll the online and target Q-networks on the sequences.
      q_values, _ = self._network.unroll(observations, core_state,
                                         self._sequence_length)
      target_q_values, _ = self._target_network.unroll(observations,
                                                       target_core_state,
                                                       self._sequence_length)

      # Compute the target policy distribution (greedy).
      greedy_actions = tf.argmax(q_values, output_type=tf.int32, axis=-1)
      target_policy_probs = tf.one_hot(
          greedy_actions, depth=self._num_actions, dtype=q_values.dtype)

      # Compute the transformed n-step loss.
      rewards = tree.map_structure(lambda x: x[:-1], rewards)
      discounts = tree.map_structure(lambda x: x[:-1], discounts)
      loss, extra = losses.transformed_n_step_loss(
          qs=q_values,
          targnet_qs=target_q_values,
          actions=actions,
          rewards=rewards,
          pcontinues=discounts * self._discount,
          target_policy_probs=target_policy_probs,
          bootstrap_n=self._n_step,
      )

      # Calculate importance weights and use them to scale the loss.
      sample_info = sample.info
      keys, probs = sample_info.key, sample_info.probability
      probs = tf2_utils.batch_to_sequence(probs)
      importance_weights = 1. / (self._max_replay_size * probs)  # [T, B]
      importance_weights **= self._importance_sampling_exponent
      importance_weights /= tf.reduce_max(importance_weights)
      loss *= tf.cast(importance_weights, tf.float32)  # [T, B]
      loss = tf.reduce_mean(loss)  # []

    # Apply gradients via optimizer.
    gradients = tape.gradient(loss, self._network.trainable_variables)
    # Clip and apply gradients.
    if self._clip_grad_norm is not None:
      gradients, _ = tf.clip_by_global_norm(gradients, self._clip_grad_norm)

    self._optimizer.apply(gradients, self._network.trainable_variables)

    # Periodically update the target network.
    if tf.math.mod(self._num_steps, self._target_update_period) == 0:
      for src, dest in zip(self._network.variables,
                           self._target_network.variables):
        dest.assign(src)
    self._num_steps.assign_add(1)

    if self._reverb_client:
      # Compute updated priorities.
      priorities = compute_priority(extra.errors, self._max_priority_weight)
      # Compute priorities and add an op to update them on the reverb side.
      self._reverb_client.update_priorities(
          table=adders.DEFAULT_PRIORITY_TABLE,
          keys=keys[:, 0],
          priorities=tf.cast(priorities, tf.float64))

    return {'loss': loss}

  def step(self):
    # Run the learning step.
    results = self._step()

    # Compute elapsed time.
    timestamp = time.time()
    elapsed_time = timestamp - self._timestamp if self._timestamp else 0
    self._timestamp = timestamp

    # Update our counts and record it.
    counts = self._counter.increment(steps=1, walltime=elapsed_time)
    results.update(counts)
    self._logger.write(results)

  def get_variables(self, names: List[str]) -> List[Variables]:
    return [tf2_utils.to_numpy(self._variables)]

  @property
  def state(self) -> Mapping[str, tf2_savers.Checkpointable]:
    """Returns the stateful parts of the learner for checkpointing."""
    return {
        'network': self._network,
        'target_network': self._target_network,
        'optimizer': self._optimizer,
        'num_steps': self._num_steps,
    }


def compute_priority(errors: tf.Tensor, alpha: float):
  """Compute priority as mixture of max and mean sequence errors."""
  abs_errors = tf.abs(errors)
  mean_priority = tf.reduce_mean(abs_errors, axis=0)
  max_priority = tf.reduce_max(abs_errors, axis=0)

  return alpha * max_priority + (1 - alpha) * mean_priority

"""Recurrent DQN (R2D2) agent implementation."""


from acme.agents.tf import actors
from acme.tf import savers as tf2_savers


class R2D22(agent.Agent):
  """R2D22 Agent.
  This implements a single-process R2D2 agent. This is a Q-learning algorithm
  that generates data via a (epislon-greedy) behavior policy, inserts
  trajectories into a replay buffer, and periodically updates the policy (and
  as a result the behavior) by sampling from this buffer.
  """

  def __init__(
      self,
      environment_spec: specs.EnvironmentSpec,
      network: snt.RNNCore,
      burn_in_length: int,
      trace_length: int,
      replay_period: int,
      counter: counting.Counter = None,
      logger: loggers.Logger = None,
      discount: float = 0.9,
      batch_size: int = 32,
      prefetch_size: int = tf.data.experimental.AUTOTUNE,
      target_update_period: int = 100,
      importance_sampling_exponent: float = 0.2,
      priority_exponent: float = 0.6,
      epsilon: float = 0.01,
      learning_rate: float = 1e-3,
      min_replay_size: int = 1000,
      max_replay_size: int = 1000000,
      samples_per_insert: float = 32.0,
      store_lstm_state: bool = True,
      max_priority_weight: float = 0.9,
      checkpoint: bool = True,
  ):

    if store_lstm_state:
      extra_spec = {
          'core_state': tf2_utils.squeeze_batch_dim(network.initial_state(1)),
      }
    else:
      extra_spec = ()

    replay_table = reverb.Table(
        name=adders.DEFAULT_PRIORITY_TABLE,
        sampler=reverb.selectors.Prioritized(priority_exponent),
        remover=reverb.selectors.Fifo(),
        max_size=max_replay_size,
        rate_limiter=reverb.rate_limiters.MinSize(min_size_to_sample=1),
        signature=adders.SequenceAdder.signature(environment_spec,
                                                   extra_spec))
    self._server = reverb.Server([replay_table], port=None)
    address = f'localhost:{self._server.port}'

    sequence_length = burn_in_length + trace_length + 1
    # Component to add things into replay.
    adder = adders.SequenceAdder(
        client=reverb.Client(address),
        period=replay_period,
        sequence_length=sequence_length,
    )

    # The dataset object to learn from.
    dataset = datasets.make_reverb_dataset(
        server_address=address,
        batch_size=batch_size,
        prefetch_size=prefetch_size,
        sequence_length=sequence_length)

    target_network = copy.deepcopy(network)
    tf2_utils.create_variables(network, [environment_spec.observations])
    tf2_utils.create_variables(target_network, [environment_spec.observations])

    learner = R2D2Learner2(
        environment_spec=environment_spec,
        network=network,
        target_network=target_network,
        burn_in_length=burn_in_length,
        sequence_length=sequence_length,
        dataset=dataset,
        reverb_client=reverb.TFClient(address),
        counter=counter,
        logger=logger,
        discount=discount,
        target_update_period=target_update_period,
        importance_sampling_exponent=importance_sampling_exponent,
        max_replay_size=max_replay_size,
        learning_rate=learning_rate,
        store_lstm_state=store_lstm_state,
        max_priority_weight=max_priority_weight,
        clip_grad_norm = 100.,
    )

    self._checkpointer = tf2_savers.Checkpointer(
        subdirectory='r2d2_learner',
        time_delta_minutes=60,
        objects_to_save=learner.state,
        enable_checkpointing=checkpoint,
    )
    self._snapshotter = tf2_savers.Snapshotter(
        objects_to_save={'network': network}, time_delta_minutes=60.)

    policy_network = snt.DeepRNN([
        network,
    ])

    actor = RecurrentActor2(
        policy_network, adder, store_recurrent_state=store_lstm_state)
    observations_per_step = (
        float(replay_period * batch_size) / samples_per_insert)
    super().__init__(
        actor=actor,
        learner=learner,
        min_observations=replay_period * max(batch_size, min_replay_size),
        observations_per_step=observations_per_step)

  def update(self):
    super().update()
    self._snapshotter.save()
    self._checkpointer.save()

# Set up a virtual display for rendering.
display = pyvirtualdisplay.Display(visible=0, size=(1400, 900)).start()


# new librariers
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px

environment = gym.make('stowkoski-env-v0')
environment = wrappers.GymWrapper(environment)  # To dm_env interface.
# Make sure the environment outputs single-precision floats.
environment = wrappers.SinglePrecisionWrapper(environment)

# Grab the spec of the environment.
environment_spec = specs.make_environment_spec(environment)

#num_dimensions = np.prod(environment_spec.actions.shape, dtype=int)
num_dimensions = 1+(17*160)#np.prod(environment_spec.actions.shape, dtype=int)

"""
#Uncomment for DQN

from acme.tf.networks.atari import  DQNAtariNetwork 
network = DQNAtariNetwork(num_dimensions)

#def _make_network(num_dim: int) -> snt.Module:
#  return snt.Sequential([
#      snt.Conv2D(32, [8, 8], [4, 4]),
#      tf.nn.relu,
#      snt.Conv2D(64, [4, 4], [2, 2]),
#      tf.nn.relu,
#      snt.Conv2D(64, [3, 3], [1, 1]),
#      tf.nn.relu,
#      snt.Flatten(),
#      snt.nets.MLP([50, 50, num_dim]),
#  ])

#network=_make_network(num_dimensions) 


agent_logger_dqn = loggers.TerminalLogger(label='agent', time_delta=1.)
env_loop_logger_dqn = loggers.TerminalLogger(label='env_loop', time_delta=1.)
csv_logger_for_agent_dqn = loggers.CSVLogger("/content/latest_agent_logger_dqn", label="agent_logging") # Logging csv file directly to colab
csv_logger_for_env_dqn = loggers.CSVLogger("/content/latest_env_logger_dqn", label="env_logging")# Logging csv file directly to colab


agent = AnotherDQN(
    environment_spec=environment_spec, 
    network=network,
    logger = csv_logger_for_agent_dqn, 
    epsilon = 0.05
    )



# Create an loop connecting this agent to the environment created above.
env_loop = EnvironmentLoop2(
    environment, agent, logger=csv_logger_for_env_dqn) # csv_logger_for_env for env does not work
#Comment till here for not DQN

"""
#Uncomment for R2D2
class SimpleNetwork(networks.RNNCore):

  def __init__(self, action_spec):
    super().__init__(name='r2d2_test_network')
    self._net = snt.DeepRNN([
        snt.Conv2D(32, [8, 8], [4, 4]),
        tf.nn.relu,
        snt.Conv2D(64, [4, 4], [2, 2]),
        tf.nn.relu,
        snt.Conv2D(64, [3, 3], [1, 1]),
        tf.nn.relu,
        snt.Flatten(),
        snt.LSTM(20),
        tf.nn.relu,
        #snt.LSTM(160),
        #snt.nets.MLP([50, 50,512]),
        #tf.nn.relu,
        snt.nets.MLP([50, 50, action_spec])
    ])

  def __call__(self, inputs, state):
    return self._net(inputs, state)

  def initial_state(self, batch_size: int, **kwargs):
    return self._net.initial_state(batch_size)

  def unroll(self, inputs, state, sequence_length):
    return snt.static_unroll(self._net, inputs, state, sequence_length)

a = SimpleNetwork(num_dimensions)
#from acme.tf.networks.atari import  R2D2AtariNetwork
#a = R2D2AtariNetwork(num_dimensions)

agent_logger_r2d2 = loggers.TerminalLogger(label='agent', time_delta=5.)
env_loop_logger_r2d2 = loggers.TerminalLogger(label='env_loop', time_delta=5.)

csv_logger_for_agent_r2d2 = loggers.CSVLogger("/content/gdrive/MyDrive/ResultsR2D2/", label="agent") # Logging csv file directly to colab
csv_logger_for_env_r2d2 = loggers.CSVLogger("/content/gdrive/MyDrive/ResultsR2D2/", label="env", time_delta=1.)# Logging csv file directly to colab


agent = R2D22(
    environment_spec=environment_spec,
    network=a,
    store_lstm_state=True,
    burn_in_length=2,
    trace_length=6,
    replay_period=4,
    checkpoint=False,
    logger = csv_logger_for_agent_r2d2
)

env_loop = EnvironmentLoop2(
    environment, agent, logger=csv_logger_for_env_r2d2) 
#Comment till here for not R2D2


env_loop.run(num_episodes=1000)

# https://github.com/deepmind/acme/blob/master/acme/environment_loop.py
# https://github.com/deepmind/acme/blob/master/acme/utils/loggers/csv.py
"""
import numpy as np
import pandas as pd
from tabulate import tabulate
import copy
import gym
import datetime
import openpyxl
import os
import xlsxwriter
from envs.environment import Environment
from stable_baselines.common.vec_env import DummyVecEnv
from stable_baselines.deepq.policies import MlpPolicy
from stable_baselines import DQN
import envs

#To open and save the results i.e makespan,machine span,state span so on to a workbook 
if os.path.isfile('makespan.xlsx'):
    os.remove('makespan.xlsx')
    wb = openpyxl.Workbook()
    wb.save('makespan.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('makespan.xlsx')
    
if os.path.isfile('machinespan.xlsx'):
    os.remove('machinespan.xlsx')
    wb = openpyxl.Workbook()
    wb.save('machinespan.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('machinespan.xlsx')
    
if os.path.isfile('stagespan.xlsx'):
    os.remove('stagespan.xlsx')
    wb = openpyxl.Workbook()
    wb.save('stagespan.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('stagespan.xlsx')
    
if os.path.isfile('jobs.xlsx'):
    os.remove('jobs.xlsx')
    wb = openpyxl.Workbook()
    wb.save('jobs.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('jobs.xlsx')
    
if os.path.isfile('output.xlsx'):
    os.remove('output.xlsx')
    wb = openpyxl.Workbook()
    wb.save('output.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('output.xlsx')
    
if os.path.isfile('machines_output.xlsx'):
    os.remove('machines_output.xlsx')
    wb = openpyxl.Workbook()
    wb.save('machines_output.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('machines_output.xlsx')
    
if os.path.isfile('contingency.xlsx'):
    os.remove('contingency.xlsx')
    wb = openpyxl.Workbook()
    wb.save('contingency.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('contingency.xlsx')
    
tstart=datetime.datetime.now()
print(datetime.datetime.now())

# data_input = pd.read_csv("input.csv",sep=",")
data_machines = pd.read_csv("machines.csv",sep=",")
# ##print(data_input)
##print(data_machines)

xls = pd.ExcelFile('P1.xlsx')

# Now you can list all sheets in the file
xls.sheet_names


sheet_to_df_map = {}

#Run the loop till all sheets of workbook are executed 
for sheet_name in xls.sheet_names:
    readData = pd.read_excel("P1.xlsx", sheet_name=sheet_name)


    #readData=readData.sort_values(by=['Start Time '])
    
    #idList=readData['Order ID'].to_list()
    
    
    sheetData= readData[['Job ID','Priority'
                         ,'Product type','Processing time (SMD)','Processing time (AOI)','Processing time (SS)',
                         'Processing time (CC)','Deadline in (min)']].copy()
    sheetData.columns=['Jobs','priority','familyType','S-1','S-2','S-3','S-4','dueDate']
    
    sheetData=sheetData.fillna(0)
   
    #sheetData = sheetData.head(20)
    #print(sheetData[sheetData['Jobs']==2])
    #sheetData.dropna(inplace=True)
    ##print(sheetData)
    
    
    
    iter_counter = 1
    env= gym.make("stowkoski-env-v0",**{"input":sheetData,"machines":data_machines,"sheet":sheet_name})
    #env =DummyVecEnv([lambda: Environment(sheetData,data_machines,sheet_name)])
    #env= Environment(sheetData,data_machines,sheet_name)
    isComplete=env.isComplete
    counter=1
    #check_env(env, True)
    #model = DQN(MlpPolicy, env, verbose=1)
    #model.learn(total_timesteps=10)
    
    obs = env.reset()
    #Run the loop till each row of the sheet gets processed
    while isComplete is not True:
        ###print(env.getAvailMachines(counter))
        ind=0
     
        avail_machines=env.getAvailMachines(counter)
    
        for i in avail_machines:
            #some_action, _states = model.predict(obs, deterministic=True)
            some_action=  env.getAction(env.getActionTime(),avail_machines)
            print(some_action)
            if some_action>0:
                #random_choice= some_action.sample()
                ###print("Random Choice :: "+str(random_choice))
                fnState,reward, isComplete,info=env.step(some_action) #some_action.iloc[0]
                ind+=1
                env.setReward()
                fixedActionMask=env.getFixedAction()
                #idList.pop(0)
                #print("job :: "+ str(tabulate(job, headers='keys', tablefmt='psql')))
                ##print("Contingency :: "+ str(tabulate(contingency, headers='keys', tablefmt='psql')))
                ##print("machine :: "+ str(tabulate(machine, headers='keys', tablefmt='psql')))
                ##print("Timetable :: "+ str(tabulate(timetable, headers='keys', tablefmt='psql')))
                #fnState.to_csv("fnstate.csv")
                ##print("State :: "+ str(tabulate(fnState, headers='keys', tablefmt='psql')))
                 
        if len(avail_machines)<=0:
            env.checkComplete()
            isComplete=env.isComplete
            if isComplete is not True:
                counter+=1
                env.updateTime()
        else:
            counter+=1
            env.updateTime()
            isComplete=env.isComplete
        iter_counter += 1
        if iter_counter%100 == 0:
            print(iter_counter)
            #timetable.to_csv("output.csv")
            print("Timetable :: "+ str(tabulate(info["timetable"], headers='keys', tablefmt='psql')))
     
    print("Game Over (::)")
    tend=datetime.datetime.now()
    print(datetime.datetime.now())
    
    print("total time :: "+str(tend-tstart))
    outputdata= pd.read_excel("output.xlsx",sheet_name=sheet_name)
    del outputdata['Unnamed: 0']
    
    out_time=outputdata.groupby('jobs').agg(['min', 'max'])['time']
    out_time['diff']=out_time['max']-out_time['min']
    
    
    
    stage_time=outputdata.groupby('stage').agg(['min', 'max'])['time']
    stage_time['diff']=stage_time['max']-stage_time['min']
    
    machine_time=outputdata.groupby('machine').agg(['min', 'max'])['time']
    machine_time['diff']=machine_time['max']-machine_time['min']
    
    
    with pd.ExcelWriter('makespan.xlsx',engine="openpyxl",mode='a') as writer1:  
        out_time.to_excel(writer1,sheet_name=sheet_name)
        writer1.close()
    with pd.ExcelWriter('stagespan.xlsx',engine="openpyxl",mode='a') as writer2:  
        stage_time.to_excel(writer2,sheet_name=sheet_name)
        writer2.close()
    with pd.ExcelWriter('machinespan.xlsx',engine="openpyxl",mode='a') as writer3:  
        machine_time.to_excel(writer3,sheet_name=sheet_name)
        writer3.close()

env= gym.make("stowkoski-env-v0")
model = DQN(MlpPolicy, env, verbose=1)
model.learn(total_timesteps=10)
"""